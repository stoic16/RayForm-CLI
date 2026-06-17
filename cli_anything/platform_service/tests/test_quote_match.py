"""Unit tests for quote_match — 产品报价核心链路 quote match.

覆盖 identify_columns / translate_vehicle_to_oe(mock翻译表) /
classify_results / build_template_excel / write_output_excel /
load_client_input(.txt) / CLI --help。

不依赖真实网络: backend 调用全部 mock，第三方查询走 deep=False 降级路径。
"""

import json
import os
import tempfile
from unittest import mock

import openpyxl
from click.testing import CliRunner

# 先导入 platform_service_cli 完成模块初始化，避免 data_clean.backend_api 的循环导入
import cli_anything.platform_service.platform_service_cli  # noqa: F401

from cli_anything.platform_service.core.data_clean.quote_match import (
    load_client_input,
    identify_columns,
    translate_vehicle_to_oe,
    build_template_excel,
    submit_and_fetch,
    classify_results,
    third_party_requery,
    write_output_excel,
    quote,
)


# ── load_client_input ──────────────────────────────────────────────────

class TestLoadClientInputTxt:
    def test_txt_each_line_is_one_row(self):
        with tempfile.NamedTemporaryFile(mode="w", suffix=".txt", delete=False,
                                          encoding="utf-8") as f:
            f.write("31110-RAA-A01\n")
            f.write("DAC3984037\n")
            f.write("  \n")  # 空行应被忽略
            f.write("本田雅阁2.4L\n")
            path = f.name

        try:
            result = load_client_input(path)
        finally:
            os.remove(path)

        assert result["headers"] is None
        assert len(result["rows"]) == 3
        assert result["rows"][0]["raw_row"] == ["31110-RAA-A01"]
        assert result["rows"][0]["row_index"] == 0
        assert result["rows"][2]["raw_row"] == ["本田雅阁2.4L"]


class TestLoadClientInputExcel:
    def test_xlsx_returns_headers_and_rows(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["工厂型号", "关联编号", "车型", "名称"])
        ws.append(["31110-RAA-A01", "31110-RBA-004,31110-RNA-A01", "本田雅阁2.4L", "轮毂轴承"])
        ws.append(["DAC3984037", "", "", ""])

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        wb.save(path)

        try:
            result = load_client_input(path)
        finally:
            os.remove(path)

        assert result["headers"] == ["工厂型号", "关联编号", "车型", "名称"]
        assert len(result["rows"]) == 2
        assert result["rows"][0]["raw_row"][0] == "31110-RAA-A01"
        assert result["rows"][0]["row_index"] == 0
        assert result["rows"][1]["raw_row"][0] == "DAC3984037"


# ── identify_columns ──────────────────────────────────────────────────

class TestIdentifyColumns:
    def test_identifies_oe_related_vehicle_columns_by_content(self):
        headers = ["工厂型号", "关联编号", "车型", "名称"]
        rows = [
            ["31110-RAA-A01", "31110-RBA-004,31110-RNA-A01", "本田雅阁2.4L", "轮毂轴承"],
            ["DAC3984037", "31110-R40-A01,31110-5T0-A01", "思域1.8L", "轴承单元"],
        ]

        cols = identify_columns(rows, headers)

        # OE 列: 单一编号，无逗号
        assert cols["oe_col"] == 0
        # 关联编号列: 含逗号分隔的多个编号
        assert cols["related_col"] == 1
        # 车型列: 含中文+数字的车型描述
        assert cols["vehicle_col"] == 2
        # 名称列: 中文产品名称
        assert cols["name_col"] == 3

    def test_no_headers_falls_back_to_content_only(self):
        rows = [
            ["31110-RAA-A01"],
            ["DAC3984037"],
            ["本田雅阁2.4L"],
        ]
        cols = identify_columns(rows, None)
        # 至少应识别出一个有效列用于后续处理，不抛异常
        assert isinstance(cols, dict)
        assert "oe_col" in cols


# ── translate_vehicle_to_oe ──────────────────────────────────────────────

FAKE_TRANSLATION_TABLE = """# 行话翻译库

## 本田

| 行话 | 发动机 | OE号 | 防御备注 |
|------|--------|------|---------|
| 本田雅阁2.4L | K24Z (八代2008-2012) | 31110-RAA-A01, 31110-RBA-004 | 不适用九代2013+款 |
| 09飞度思迪1.5L | L15A | 31110-PNA-006 | 不适用1.3L |

## 吉利

| 行话 | 发动机 | OE号 | 防御备注 |
|------|--------|------|---------|
| 吉利博越/EC8/GC7/GX7/远景SUV 2.0/2.4L | 吉利自主 | — | ⚠ 吉利自主品牌，17vin/泰安联均无覆盖 |
"""


class TestTranslateVehicleToOe:
    def test_found_in_table_returns_first_oe(self):
        with mock.patch(
            "cli_anything.platform_service.core.data_clean.quote_match._load_translation_table",
            return_value=FAKE_TRANSLATION_TABLE,
        ):
            oe = translate_vehicle_to_oe("本田雅阁2.4L", deep=False)

        assert oe == "31110-RAA-A01"

    def test_table_entry_with_no_oe_returns_none(self):
        with mock.patch(
            "cli_anything.platform_service.core.data_clean.quote_match._load_translation_table",
            return_value=FAKE_TRANSLATION_TABLE,
        ):
            oe = translate_vehicle_to_oe("吉利博越", deep=False)

        assert oe is None

    def test_not_found_and_not_deep_returns_none_without_network(self):
        with mock.patch(
            "cli_anything.platform_service.core.data_clean.quote_match._load_translation_table",
            return_value=FAKE_TRANSLATION_TABLE,
        ) as mock_table, mock.patch(
            "cli_anything.platform_service.core.data_clean.quote_match.search_17vin"
        ) as mock_17vin, mock.patch(
            "cli_anything.platform_service.core.data_clean.quote_match.search_tecalliance"
        ) as mock_tec:
            oe = translate_vehicle_to_oe("未知车型XYZ", deep=False)

        assert oe is None
        mock_17vin.assert_not_called()
        mock_tec.assert_not_called()

    def test_not_found_and_deep_but_cdp_unreachable_degrades_to_none(self):
        with mock.patch(
            "cli_anything.platform_service.core.data_clean.quote_match._load_translation_table",
            return_value=FAKE_TRANSLATION_TABLE,
        ), mock.patch(
            "cli_anything.platform_service.core.data_clean.quote_match.check_cdp_ready",
            return_value=False,
        ), mock.patch(
            "cli_anything.platform_service.core.data_clean.quote_match.search_17vin"
        ) as mock_17vin:
            oe = translate_vehicle_to_oe("未知车型XYZ", deep=True)

        assert oe is None


# ── classify_results ──────────────────────────────────────────────────

class TestClassifyResults:
    def _row(self, price_check_value, query_source, product_id="p1", **extra):
        row = {
            "priceCheckValue": price_check_value,
            "querySource": query_source,
            "productId": product_id,
            "code": extra.get("code", ""),
            "name": extra.get("name", ""),
        }
        row.update(extra)
        return row

    def test_unique_exact_match_goes_to_matched(self):
        rows = [self._row("31110-RAA-A01", 1)]
        result = classify_results(rows)

        assert len(result["matched"]) == 1
        assert result["matched"][0]["priceCheckValue"] == "31110-RAA-A01"
        assert result["to_review"] == []
        assert result["unmatched"] == []

    def test_fuzzy_match_goes_to_to_review(self):
        rows = [self._row("31110-XXXXX", 5)]
        result = classify_results(rows)

        assert result["matched"] == []
        assert len(result["to_review"]) == 1
        assert result["unmatched"] == []

    def test_duplicate_price_check_value_goes_to_to_review(self):
        rows = [
            self._row("31110-RAA-A01", 1, product_id="p1"),
            self._row("31110-RAA-A01", 4, product_id="p2"),
        ]
        result = classify_results(rows)

        assert result["matched"] == []
        assert len(result["to_review"]) == 2
        assert result["unmatched"] == []

    def test_no_result_goes_to_unmatched(self):
        rows = [self._row("UNKNOWN-CODE-999", 0)]
        result = classify_results(rows)

        assert result["matched"] == []
        assert result["to_review"] == []
        assert len(result["unmatched"]) == 1

    def test_mixed_rows_classified_correctly(self):
        rows = [
            self._row("31110-RAA-A01", 1),                       # matched
            self._row("DAC3984037", 4, product_id="p2"),          # matched
            self._row("31110-DUP", 2, product_id="p3"),           # to_review (dup)
            self._row("31110-DUP", 6, product_id="p4"),           # to_review (dup)
            self._row("FUZZY-OE", 5, product_id="p5"),            # to_review (fuzzy)
            self._row("NOT-FOUND", 0, product_id=""),             # unmatched
        ]
        result = classify_results(rows)

        matched_values = {r["priceCheckValue"] for r in result["matched"]}
        review_values = {r["priceCheckValue"] for r in result["to_review"]}
        unmatched_values = {r["priceCheckValue"] for r in result["unmatched"]}

        assert matched_values == {"31110-RAA-A01", "DAC3984037"}
        assert review_values == {"31110-DUP", "FUZZY-OE"}
        assert unmatched_values == {"NOT-FOUND"}
        assert len(result["to_review"]) == 3


# ── build_template_excel ──────────────────────────────────────────────

class TestBuildTemplateExcel:
    def test_generates_excel_with_required_headers(self):
        rows = [
            {"OE": "31110-RAA-A01", "通用OE": "31110-RBA-004", "名称": "轮毂轴承",
             "标签车型": "本田雅阁2.4L", "通用车型": "", "销售等级": "A"},
            {"OE": "DAC3984037", "通用OE": "", "名称": "", "标签车型": "", "通用车型": "", "销售等级": ""},
        ]

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = os.path.join(tmpdir, "template.xlsx")
            build_template_excel(rows, output_path)

            assert os.path.exists(output_path)
            wb = openpyxl.load_workbook(output_path)
            ws = wb.active
            header_row = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
            assert header_row == ["OE", "通用OE", "名称", "标签车型", "通用车型", "销售等级"]

            data_row1 = [c.value for c in next(ws.iter_rows(min_row=2, max_row=2))]
            assert data_row1[0] == "31110-RAA-A01"
            assert data_row1[1] == "31110-RBA-004"


# ── submit_and_fetch ──────────────────────────────────────────────────

class TestSubmitAndFetch:
    def test_uploads_file_queries_list_and_fetches_detail(self):
        backend = mock.MagicMock()

        # parse 上传成功
        backend.post.return_value = {"code": 200, "msg": "OK", "status": True, "data": None}

        # list 查询返回最新一条审核任务
        list_resp = {
            "code": 200, "status": True,
            "data": {"content": [{"id": "audit-001", "importFileName": "template.xlsx"}],
                      "totalElements": 1},
        }

        # findAll 返回明细
        detail_resp = {
            "code": 200, "status": True,
            "data": [
                {"priceCheckValue": "31110-RAA-A01", "querySource": 1, "productId": "p1"},
            ],
        }

        backend.get.side_effect = [list_resp, detail_resp]

        with tempfile.TemporaryDirectory() as tmpdir:
            template_path = os.path.join(tmpdir, "template.xlsx")
            build_template_excel(
                [{"OE": "31110-RAA-A01", "通用OE": "", "名称": "", "标签车型": "", "通用车型": "", "销售等级": ""}],
                template_path,
            )

            audit_id, detail_rows = submit_and_fetch(
                backend, template_path,
                supplier_range="", query_range="0,1,2,3,4", query_repair_kit=False,
            )

        assert audit_id == "audit-001"
        assert detail_rows == detail_resp["data"]

        # parse 调用应包含 files
        post_args, post_kwargs = backend.post.call_args
        assert post_args[0] == "/api/principal/productAudit/parse"
        assert "files" in post_kwargs

    def test_findall_empty_object_returns_empty_list(self):
        backend = mock.MagicMock()
        backend.post.return_value = {"code": 200, "status": True, "data": None}
        backend.get.side_effect = [
            {"code": 200, "status": True,
             "data": {"content": [{"id": "audit-002", "importFileName": "template.xlsx"}]}},
            {"code": 200, "status": True, "data": {}},  # findAll 空结果
        ]

        with tempfile.TemporaryDirectory() as tmpdir:
            template_path = os.path.join(tmpdir, "template.xlsx")
            build_template_excel(
                [{"OE": "NOT-FOUND", "通用OE": "", "名称": "", "标签车型": "", "通用车型": "", "销售等级": ""}],
                template_path,
            )
            audit_id, detail_rows = submit_and_fetch(
                backend, template_path,
                supplier_range="", query_range="0,1,2,3,4", query_repair_kit=False,
            )

        assert audit_id == "audit-002"
        assert detail_rows == []


# ── third_party_requery ──────────────────────────────────────────────

class TestThirdPartyRequery:
    def test_no_deep_degrades_all_to_still_unmatched(self):
        unmatched = [{"priceCheckValue": "UNKNOWN-1", "querySource": 0, "productId": ""}]

        requery_matched, still_unmatched = third_party_requery(unmatched, deep=False)

        assert requery_matched == []
        assert still_unmatched == unmatched

    def test_deep_but_cdp_unreachable_degrades_without_crash(self):
        unmatched = [{"priceCheckValue": "UNKNOWN-1", "querySource": 0, "productId": ""}]

        with mock.patch(
            "cli_anything.platform_service.core.data_clean.quote_match.check_cdp_ready",
            return_value=False,
        ):
            requery_matched, still_unmatched = third_party_requery(unmatched, deep=True)

        assert requery_matched == []
        assert still_unmatched == unmatched

    def test_deep_cdp_ready_but_search_raises_degrades_gracefully(self):
        unmatched = [{"priceCheckValue": "UNKNOWN-1", "querySource": 0, "productId": ""}]

        with mock.patch(
            "cli_anything.platform_service.core.data_clean.quote_match.check_cdp_ready",
            return_value=True,
        ), mock.patch(
            "cli_anything.platform_service.core.data_clean.quote_match.search_17vin",
            side_effect=RuntimeError("network error"),
        ):
            requery_matched, still_unmatched = third_party_requery(unmatched, deep=True)

        assert requery_matched == []
        assert still_unmatched == unmatched


# ── write_output_excel ──────────────────────────────────────────────

class TestWriteOutputExcel:
    def test_writes_four_sheets(self):
        matched = [{
            "priceCheckValue": "31110-RAA-A01", "productId": "p1", "name": "轮毂轴承",
            "code": "RAH3C43", "querySource": 1, "salePrice": 100, "oemPrice": 90,
            "suggestPrice": 110, "p1Price": 95, "p2Price": 98, "p3Price": 102,
            "count": 5,
        }]
        to_review = [
            {"priceCheckValue": "31110-DUP", "productId": "p2", "name": "轴承A",
             "code": "C1", "querySource": 2, "salePrice": 50, "oemPrice": 45,
             "suggestPrice": 55, "p1Price": 48, "p2Price": 49, "p3Price": 52, "count": 3},
            {"priceCheckValue": "31110-DUP", "productId": "p3", "name": "轴承B",
             "code": "C2", "querySource": 6, "salePrice": 60, "oemPrice": 55,
             "suggestPrice": 65, "p1Price": 58, "p2Price": 59, "p3Price": 62, "count": 2},
        ]
        requery_matched = [{
            "priceCheckValue": "OLD-OE-1", "productId": "p4", "name": "轴承C",
            "code": "C3", "querySource": 1, "salePrice": 70, "oemPrice": 65,
            "suggestPrice": 75, "p1Price": 68, "p2Price": 69, "p3Price": 72,
            "count": 1, "new_oe": "NEW-OE-1", "source": "17vin",
            "original_input": "OLD-OE-1",
        }]
        still_unmatched = [{"priceCheckValue": "UNKNOWN-9", "querySource": 0, "productId": ""}]

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = os.path.join(tmpdir, "output.xlsx")
            write_output_excel(matched, to_review, requery_matched, still_unmatched, output_path)

            assert os.path.exists(output_path)
            wb = openpyxl.load_workbook(output_path)
            assert wb.sheetnames == ["报价结果", "待技术员分辨", "三方补查待写入", "待工厂确认"]

            # 报价结果 sheet 应包含匹配行
            ws1 = wb["报价结果"]
            rows1 = list(ws1.iter_rows(min_row=2, values_only=True))
            assert len(rows1) == 1
            assert rows1[0][0] == "31110-RAA-A01"

            # 待技术员分辨 sheet 应包含两条重复行
            ws2 = wb["待技术员分辨"]
            rows2 = list(ws2.iter_rows(min_row=2, values_only=True))
            assert len(rows2) == 2

            # 三方补查待写入 sheet 应包含新OE/来源列
            ws3 = wb["三方补查待写入"]
            header3 = [c.value for c in next(ws3.iter_rows(min_row=1, max_row=1))]
            assert "新OE" in header3
            assert "来源" in header3

            # 待工厂确认 sheet
            ws4 = wb["待工厂确认"]
            rows4 = list(ws4.iter_rows(min_row=2, values_only=True))
            assert len(rows4) == 1


# ── CLI smoke test ──────────────────────────────────────────────────

class TestQuoteCli:
    def test_quote_match_help(self):
        runner = CliRunner()
        result = runner.invoke(quote, ["match", "--help"])

        assert result.exit_code == 0
        assert "match" in result.output or "--file" in result.output
